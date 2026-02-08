# Gaya
import ipaddress
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
 
# VLAN definitions
vlans = [
    ("Camera_Pool", 10),
    ("ATCS_ECB_Pool", 11),
    ("UPS_SW_Pool", 12),
    ("VMD_Radar_Pool", 13)
]
 
locations = [
    "Girja Chowk",
    "Janta Chowk",
    "Max Hospital Tri chowk",
    "Bihar Takieej Road",
    "Bus Stand",
    "Polytechnic Chowk",
    "Mata Sathaan Chowk",
    "Vishveshriya Chowk",
    "R. N Saah Chowk",
    "Line Bazar Chowk",
    "Katihaar Chowk",
    "Madhubani Chowk",
    "Manjhaali Chowk",
    "Dolar house Chowk",
    "Rambagh Chowk",
    "Khushkibagh",
    "Station Road",
    "Rajni Chowk",
    "Jhanda Chowk/Bhatha Bazaar",
    "Kheru chowk",
    "Thanna k Pass",
    "Jeella School Road",
    "Maranga Chowk",
    "DIG Chowk",
    "Govt. Medical Collage and Hospital",
    "Near Civil Court Purnia",
    "Near Power grid 1",
    "Sudin Chowk Market",
    "Near Rizwan Masjid",
    "Near Jail Chowk",
    "Near Polytechnic Chowk",
    "Near SBI Main branch mod",
    "Near Mahakali motors",
    "Janta Chowk more Rambagh",
    "Near power grid 2",
    "Near Kali Mandir Ashram Rd",
    "Nagar Nigam Chowk",
    "Maffia Chowk Near Petrol pump",
    "Near DAV School",
    "RangBhoomi Chowk",
    "Rahman Chowk",
    "Near CPWD Office",
    "Milky Chowk (Near Purnia Semen Station)",
    "Zero Mile",
    "ICCC"
]

allocation_counts = [
    6, 2, 2, 2, 2, 2, 2, 4, 6, 6, 5, 2, 2, 2, 2, 2, 2, 2, 2, 2, 3, 2, 5, 2, 2, 2, 2, 2, 3, 2, 2, 2, 2, 2, 3, 2, 3, 2, 3, 2, 2, 2, 2, 2, 3
]
 
 
 
# Starting subnet for each VLAN block
current_subnet = ipaddress.ip_network("10.10.128.0/26")
 
# Create workbook
wb = Workbook()
ws = wb.active
ws.title = "IP Plan"
 
# Headers
headers = [
    "S.No", "Location Name",
    "Nos of 8-Port SW", "Usable IPs",
    "VLAN Name", "VLAN ID",
    "IP Subnet"
]
ws.append(headers)
 
# Styling
header_fill = PatternFill("solid", fgColor="FFFF00")
border = Border(left=Side(style="thin"), right=Side(style="thin"),
                top=Side(style="thin"), bottom=Side(style="thin"))
center = Alignment(horizontal="center", vertical="center")
 
# Apply header formatting
for col in range(1, len(headers) + 1):
    cell = ws.cell(row=1, column=col)
    cell.fill = header_fill
    cell.border = border
    cell.font = Font(bold=True)
    cell.alignment = center
 
serial = 1
 
# Fill data
for idx, location in enumerate(locations):
    sw_count = allocation_counts[idx]
    usable_ips = sw_count * 8
 
    first_row = ws.max_row + 1
 
    for vlan_name, vlan_id in vlans:
        ws.append([
            serial,
            location,
            sw_count,
            usable_ips,
            vlan_name,
            vlan_id,
            str(current_subnet)
        ])
 
        # Move to next subnet
        current_subnet = ipaddress.ip_network(
            f"{current_subnet.network_address + current_subnet.num_addresses}/26"
        )
 
    # Merge S.No
    ws.merge_cells(start_row=first_row, start_column=1, end_row=first_row + 3, end_column=1)
    ws.cell(first_row, 1).alignment = center
 
    # Merge Location
    ws.merge_cells(start_row=first_row, start_column=2, end_row=first_row + 3, end_column=2)
    ws.cell(first_row, 2).alignment = center
 
    # Merge switch count
    ws.merge_cells(start_row=first_row, start_column=3, end_row=first_row + 3, end_column=3)
    ws.cell(first_row, 3).alignment = center
 
    # Merge usable IP
    ws.merge_cells(start_row=first_row, start_column=4, end_row=first_row + 3, end_column=4)
    ws.cell(first_row, 4).alignment = center
 
    serial += 1
 
# Apply borders everywhere
for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=1, max_col=len(headers)):
    for cell in row:
        cell.border = border
 
# Auto column width
for column_cells in ws.columns:
    length = max(len(str(cell.value)) for cell in column_cells)
    ws.column_dimensions[column_cells[0].column_letter].width = length + 5
 
# Save file
wb.save("Traffic_IP_Plan_With_8PortSW.xlsx")
 
print("Excel Generated Successfully: Traffic_IP_Plan_With_8PortSW.xlsx")
 
 
 