
# # GAYA – 75 Locations – HORIZONTAL + EMPTY COLUMN BETWEEN VLANs – FULL 62 IPs
# import ipaddress
# from openpyxl import Workbook
# from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
# from openpyxl.utils import get_column_letter
 
# # =============================================
# # 1. DATA (75 locations + allocations)
# # =============================================
# locations = [
#     "Kasinath Mod", "Kacheri ke samne", "D M Golambar", "Peer Mansur Chowk",
#     "Chata Masjid ke samne GB Road", "Nadraganj Masjid ke samne", "Samir Takiya Mod",
#     "MirzaGalib College Mod", "Nagmitya Mod", "Railway hospital ke pas",
#     "Karimganj Dhela Pul Daksin chor", "Kotwali Thana Mod", "Golpattar Uttari",
#     "Dukharini Mandir ke pass", "Kirani Ghat pul ke neeche", "Tower Chowk",
#     "Ramshila Mod", "Pretshila Mod (Gandhi Chowk)", "Station ke samne Sadak par",
#     "Bata More Teenmuhani", "6 lane Falgu Pul ke Uppar Pashchim Taraf",
#     "Purani Godam Shouchalya ke pas", "Kali Stan chowk", "Nai Godam Mod",
#     "Chand Chaura Purab mod", "Chad Chaura paschimi", "Manglagauri Mod",
#     "ITI ke samne Bodh Gaya Rd", "Ghungaditad Chowk", "Nawagadi Mod",
#     "Vishnupad Mandir ke samne", "Bangali Ashram Mod", "Falgu nadi ke Pul Pashchimi chor",
#     "Gewal Bigha Mod", "Jai Prakash Jharna Tinmuhani", "Sikariya Mod",
#     "SSP awas ke pass (Sudha Dairy)", "Gaya College Mod",
#     "Chandoti Mod Kendriya Kara Gaya k Dakshin Pravesh Dwaar",
#     "City SP karyalaya Tiraha", "Samir Takkiya Durga Sathan ke samne",
#     "Munni Masjid Chowk", "Chandoti Mod", "Behla thana Mod",
#     "Behla Pul uttari chor", "Medical College gate ke samne",
#     "Medical College emergency gate", "BIPARD ke samne", "05 no. gate gaya dobi road",
#     "Vishnuganj Bazar", "Katari Bajrang bali mod", "Candauti Thana ke samne",
#     "Ambedkar Chowk chandauti thana", "Bunyad Ganj Bypas(NH82 Patna Gaya Rd)",
#     "Chankad Bazar", "Bitho Sarif Mazar wali sadak", "Dhanava Mod",
#     "Tekuna param Mod", "Janpur Mod", "Do Muhan par (Charo Disha Mein)",
#     "MVV Main gate ke pass", "Mufsil Mod", "Sixlane Purvi Chor Par", "Bhusunda Mod",
#     "Mehta Petrol Pump Railway Overbridge", "Abgila Masjid ke pas",
#     "City Public School ke pas 4 lane", "Green Field School", "Nawada Bus Stand ke pas",
#     "Kijarsarai Mod", "Bypass pul purvi Chor par", "Khanjapur chowk",
#     "Kalai Teen Muhan (harijan Dharmashala)", "TOP manpur Bunyadgun Gopalpur Railway",
#     "SH04 Khijarsarai Seema"
# ]
 
# allocation_counts = [6,5,4,2,5,4,2,4,2,4,2,5,2,2,2,3,2,2,2,2,2,2,2,2,2,2,2,2,2,2,2,2,2,2,2,2,2,2,2,2,2,2,2,2,2,2,2,2,2,2,2,2,2,2,2,2,2,2,3,2,2,2,2,2,2,2,2,2,2,2,2,2,2,2,2,2,2,2,2]
 
# vlan_info = {10: "Camera_Pool", 11: "ATCS_ECB_Pool", 12: "UPS_SW_Pool", 13: "VMD_Radar_Pool"}
# colors    = {10: "CFE2F3", 11: "D9EAD3", 12: "F4CCCC", 13: "FFE599"}
 
# # =============================================
# # 2. STYLES
# # =============================================
# bold = Font(bold=True)
# center = Alignment(horizontal="center", vertical="center")
# border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
 
# # =============================================
# # 3. CREATE WORKBOOK
# # =============================================
# wb = Workbook()
# wb.remove(wb.active)
# base = ipaddress.IPv4Network("10.10.128.0/26")
# subnet_counter = 0
 
# for loc_idx, location in enumerate(locations):
#     ws = wb.create_sheet(title=location[:31])
#     switches = allocation_counts[loc_idx]
#     row = 1
 
#     # === Title ===
#     ws.merge_cells('A1:S1')  # now 19 columns
#     ws['A1'] = location
#     ws['A1'].fill = PatternFill("solid", "1F4E79")
#     ws['A1'].font = Font(color="FFFFFF", bold=True, size=14)
#     ws['A1'].alignment = center
#     row += 2
 
#     # === VLAN Headers + Empty Gap ===
#     col = 1
#     vlan_subnets = []
#     for vlan in [10, 11, 12, 13]:
#         net_int = int(base.network_address) + subnet_counter * 64
#         subnet = ipaddress.IPv4Network((net_int, 26))
#         vlan_subnets.append(subnet)
#         subnet_counter += 1
 
#         # VLAN Title (merged over 4 columns)
#         ws.merge_cells(start_row=row, start_column=col, end_row=row, end_column=col+3)
#         cell = ws.cell(row=row, column=col, value=f"VLAN {vlan} – {vlan_info[vlan]} – {subnet}")
#         cell.fill = PatternFill("solid", colors[vlan])
#         cell.font = bold
#         cell.alignment = center
 
#         # Column headers
#         headers = ["S.No", "IP Address", "Allocation", "Device"]
#         for i, h in enumerate(headers):
#             c = ws.cell(row=row+1, column=col+i, value=h)
#             c.fill = PatternFill("solid", "FFFF00")
#             c.font = bold
#             c.alignment = center
#             c.border = border
 
#         col += 5  # 4 data columns + 1 empty gap (except after last VLAN)
 
#     # Adjust last gap (no empty column after VLAN 13)
#     col -= 1
#     row += 2
 
#     # === Fill 62 IPs per VLAN ===
#     all_ips = [list(subnet.hosts()) for subnet in vlan_subnets]
 
#     for ip_idx in range(62):
#         current_row = row + ip_idx
#         col = 1
#         for vlan_idx, vlan in enumerate([10, 11, 12, 13]):
#             ip = all_ips[vlan_idx][ip_idx]
#             pool = vlan_info[vlan]
 
#             ws.cell(current_row, col, ip_idx + 1).alignment = center
#             ws.cell(current_row, col + 1, str(ip))
 
#             if ip_idx == 0:
#                 alloc = "Gateway IP"        
#             elif ip_idx < 10:
#                 alloc = "Reserved IP"
#             elif vlan == 12 and ip_idx == 61:
#                 alloc = "UPS"
#             else:
#                 alloc = pool
#             ws.cell(current_row, col + 2, alloc)
 
#             col += 5  # move to next VLAN block (skip empty column)
 
#     # === Column widths ===
#     widths = [8, 18, 18, 30, 3, 8, 18, 18, 30, 3, 8, 18, 18, 30, 3, 8, 18, 18, 30]  # 19 columns
#     for i, w in enumerate(widths, 1):
#         ws.column_dimensions[get_column_letter(i)].width = w
 
# # =============================================
# # 4. SAVE
# # =============================================
# wb.save("Gaya.xlsx")
# print("SUCCESS! File created: Gaya_75_Locations_HORIZONTAL_WITH_GAP.xlsx")
# print("   → One empty column between each VLAN block")
# print("   → All 62 usable IPs fully displayed")
 
###########################################################################################################
 
# Chapra
# # GAYA – 75 Locations – HORIZONTAL + EMPTY COLUMN BETWEEN VLANs – FULL 62 IPs
# import ipaddress
# from openpyxl import Workbook
# from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
# from openpyxl.utils import get_column_letter
 
# # =============================================
# # 1. DATA (75 locations + allocations)
# # =============================================
# locations = [
#     "Traffic Number 1",
#     "Rajiv Gandhi Chowk",
#     "Bata Chowk",
#     "Deen Dayal Chowk",
#     "Gandhi Chowk",
#     "Murgiya Chowk",
#     "Dilip Babu Mehal station",
#     "Purabhsaraya Masjid",
#     "Gaushala Chowk",
#     "Gumti no 5 tinbatiya chowk",
#     "Ambe Chowk",
#     "Nilam Chowk",
#     "Ambedkar Chowk",
#     "Kila north gate Purani Police Line",
#     "Sadar Hospital",
#     "Basudevpur OP(Chowk)",
#     "2 Number Gumti",
#     "I.T.C Park",
#     "Chandika asthaan",
#     "Shubhash Chowk",
#     "3 Number Gumti",
#     "Bhagat Singh Chowk",
#     "Sitariya Petrol Pump",
#     "Rajendra Chowk (Jubli Well)",
#     "Government Bus Stand",
#     "Sojhi Ghat",
#     "Babua Ghat",
#     "Kashtaharni Ghat",
#     "khankah Machali Mod",
#     "Chandan Bagh Thanna Mod",
#     "Koura Maidan",
#     "Dakra Nalla Mod",
#     "Telliya Talab Mod",
#     "SBI Mod Chowk",
#     "Jubliwell Chowk",
#     "Doulatpur Chowk",
#     "Panch Mukhi Mandir Mod",
#     "Station Chowk 6 number gate",
#     "Bharat Mata Chowk",
#     "Janta Chowk",
#     "Sahjuber more",
#     "lal darwaja qila",
#     "vijay chowk",
#     "Kasturba water",
#     "Badi Durga mandir Dakshin main gate",
#     "Thana chowk",
#     "Herudiyara morcha ke pass",
#     "Hajaratganj chowk",
#     "Hasanganj imali chouk",
#     "Bharat chowk",
#     "Badi Bazaar",
#     "Chua Bagh Machli Mod",
#     "East Colony Karkhana Mod",
#     "Konarke Mod",
#     "Lallu Pokhar Tinbatiya",
#     "Muffasil Thanna Chowk",
#     "DJ College",
#     "Shastri Chowk"
# ]
 
# allocation_counts = [6,5,4,2,5,4,2,4,2,4,2,5,2,2,2,3,2,2,2,2,2,2,2,2,2,2,2,2,2,2,2,2,2,2,2,2,2,2,2,2,2,2,2,2,2,2,2,2,2,2,2,2,2,2,2,2,2,2,3,2,2,2,2,2,2,2,2,2,2,2,2,2,2,2,2,2,2,2,2]
 
# vlan_info = {10: "Camera_Pool", 11: "ATCS_ECB_Pool", 12: "UPS_SW_Pool", 13: "VMD_Radar_Pool"}
# colors    = {10: "CFE2F3", 11: "D9EAD3", 12: "F4CCCC", 13: "FFE599"}
 
# # =============================================
# # 2. STYLES
# # =============================================
# bold = Font(bold=True)
# center = Alignment(horizontal="center", vertical="center")
# border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
 
# # =============================================
# # 3. CREATE WORKBOOK
# # =============================================
# wb = Workbook()
# wb.remove(wb.active)
# base = ipaddress.IPv4Network("10.10.128.0/26")
# subnet_counter = 0
 
# for loc_idx, location in enumerate(locations):
#     ws = wb.create_sheet(title=location[:31])
#     switches = allocation_counts[loc_idx]
#     row = 1
 
#     # === Title ===
#     ws.merge_cells('A1:S1')  # now 19 columns
#     ws['A1'] = location
#     ws['A1'].fill = PatternFill("solid", "1F4E79")
#     ws['A1'].font = Font(color="FFFFFF", bold=True, size=14)
#     ws['A1'].alignment = center
#     row += 2
 
#     # === VLAN Headers + Empty Gap ===
#     col = 1
#     vlan_subnets = []
#     for vlan in [10, 11, 12, 13]:
#         net_int = int(base.network_address) + subnet_counter * 64
#         subnet = ipaddress.IPv4Network((net_int, 26))
#         vlan_subnets.append(subnet)
#         subnet_counter += 1
 
#         # VLAN Title (merged over 4 columns)
#         ws.merge_cells(start_row=row, start_column=col, end_row=row, end_column=col+3)
#         cell = ws.cell(row=row, column=col, value=f"VLAN {vlan} – {vlan_info[vlan]} – {subnet}")
#         cell.fill = PatternFill("solid", colors[vlan])
#         cell.font = bold
#         cell.alignment = center
 
#         # Column headers
#         headers = ["S.No", "IP Address", "Allocation", "Device"]
#         for i, h in enumerate(headers):
#             c = ws.cell(row=row+1, column=col+i, value=h)
#             c.fill = PatternFill("solid", "FFFF00")
#             c.font = bold
#             c.alignment = center
#             c.border = border
 
#         col += 5  # 4 data columns + 1 empty gap (except after last VLAN)
 
#     # Adjust last gap (no empty column after VLAN 13)
#     col -= 1
#     row += 2
 
#     # === Fill 62 IPs per VLAN ===
#     all_ips = [list(subnet.hosts()) for subnet in vlan_subnets]
 
#     for ip_idx in range(62):
#         current_row = row + ip_idx
#         col = 1
#         for vlan_idx, vlan in enumerate([10, 11, 12, 13]):
#             ip = all_ips[vlan_idx][ip_idx]
#             pool = vlan_info[vlan]
 
#             ws.cell(current_row, col, ip_idx + 1).alignment = center
#             ws.cell(current_row, col + 1, str(ip))
 
#             if ip_idx == 0:
#                 alloc = "Gateway IP"        
#             elif ip_idx < 10:
#                 alloc = "Reserved IP"
#             elif vlan == 12 and ip_idx == 61:
#                 alloc = "UPS"
#             else:
#                 alloc = pool
#             ws.cell(current_row, col + 2, alloc)
 
#             col += 5  # move to next VLAN block (skip empty column)
 
#     # === Column widths ===
#     widths = [8, 18, 18, 30, 3, 8, 18, 18, 30, 3, 8, 18, 18, 30, 3, 8, 18, 18, 30]  # 19 columns
#     for i, w in enumerate(widths, 1):
#         ws.column_dimensions[get_column_letter(i)].width = w
 
# # =============================================
# # 4. SAVE
# # =============================================
# wb.save("Chapra.xlsx")
# print("SUCCESS! File created: Gaya_75_Locations_HORIZONTAL_WITH_GAP.xlsx")
# print("   → One empty column between each VLAN block")
# print("   → All 62 usable IPs fully displayed")
 
#########################################################################################################
# Munger
 
import ipaddress
import sys
import io

sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8')
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter
 
# =============================================
# 1. DATA (75 locations + allocations)
# =============================================
 
#  Locations
locations = [
    "Nagar Thanna Chowk",
    "Nagar Palika Chowk",
    "Yogniya Kothi Mod",
    "Sada Over Bridge at west side",
    "Sada Over Bridge ke niche at fish market tri road",
    "Mona Chowk",
    "Mewalal Chowk",
    "Gandhi Chowk",
    "Kathari Bagh tri chowk",
    "Khanua Nalla",
    "Sahebganj Chowk",
    "Nalla No 2- Sonarpatti Road",
    "Mehmood Chowk",
    "Bhikari Chowk",
    "Nehru Chowk",
    "Shishu Park main road",
    "Sadar Hospital Chowk near Mazaar",
    "Daroga Rai Chowk",
    "Bahart Milaap Chowk",
    "Bhagwaan Bazar Station Chowk",
    "Bhagwaan Bazar Thanna Chowk",
    "Near Gudari Bazar Masjid at Chaar Muhani",
    "Barahampur Chowk",
    "Shayamchak",
    "Koniyaa Mayi mandir Chowk",
    "Rajendra Sarovaar Chowk",
    "Mohaya Overbridge both side- Entry",
    "Mohaya Overbridge both side- Exit",
    "Mohaya Overbridge ke Niche (New Bypass )",
    "Novajee Tolla Chowk",
    "Bazar Samiti Chowk",
    "Methvalliya Chowk",
    "Mathiya Chowk",
    "Umdha chowk",
    "Sadda Dalla Overbridge at North side",
    "Chanchora Bazar",
    "Near Magaiedeh Railway Dalla 51",
    "Jhanga Chowk (Near Doriganj Pull)",
    "Kachehri station mod",
    "Kashi Bazar Chowk",
    "Prabhu Nath Nagar PCC mod",
    "Bishnpura Mohiya mode",
    "Mukrera mod",
    "Bada Telpa",
    "Police Line Gate",
    "Lower road nayi Bazaar mod (teen mohani)",
    "Lal Bazaar",
    "Garkha dhala rathore tola",
    "Gheghtha Bazar",
    "Meera Musheri Mod",
    "Dwarka Dheesh Mod",
    "Turkaulia",
    "Near Enayi Hanumaan Mandir",
    "State Bank of India Near Pankaj Cinema",
    "Bazar Samiti Main Gate",
    "Chota Telpa Chowk"
]
allocation_counts = [6,5,4,2,5,4,2,4,2,4,2,5,2,2,2,3,2,2,2,2,2,2,2,2,2,2,2,2,2,2,2,2,2,2,2,2,2,2,2,2,2,2,2,2,2,2,2,2,2,2,2,2,2,2,2,2,2,2,3,2,2,2,2,2,2,2,2,2,2,2,2,2,2,2,2,2,2,2,2]
 
vlan_info = {10: "Camera_Pool", 11: "ATCS_ECB_Pool", 12: "UPS_SW_Pool", 13: "VMD_Radar_Pool"}
colors    = {10: "CFE2F3", 11: "D9EAD3", 12: "F4CCCC", 13: "FFE599"}
 
# =============================================
# 2. STYLES
# =============================================
bold = Font(bold=True)
center = Alignment(horizontal="center", vertical="center")
border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
 
# =============================================
# 3. CREATE WORKBOOK
# =============================================
wb = Workbook()
wb.remove(wb.active)
base = ipaddress.IPv4Network("10.10.128.0/26")
subnet_counter = 0
 
for loc_idx, location in enumerate(locations):
    ws = wb.create_sheet(title=location[:31])
    switches = allocation_counts[loc_idx]
    row = 1
 
    # === Title ===
    ws.merge_cells('A1:S1')  # now 19 columns
    ws['A1'] = location
    ws['A1'].fill = PatternFill("solid", "1F4E79")
    ws['A1'].font = Font(color="FFFFFF", bold=True, size=14)
    ws['A1'].alignment = center
    row += 2
 
    # === VLAN Headers + Empty Gap ===
    col = 1
    vlan_subnets = []
    for vlan in [10, 11, 12, 13]:
        net_int = int(base.network_address) + subnet_counter * 64
        subnet = ipaddress.IPv4Network((net_int, 26))
        vlan_subnets.append(subnet)
        subnet_counter += 1
 
        # VLAN Title (merged over 4 columns)
        ws.merge_cells(start_row=row, start_column=col, end_row=row, end_column=col+3)
        cell = ws.cell(row=row, column=col, value=f"VLAN {vlan} – {vlan_info[vlan]} – {subnet}")
        cell.fill = PatternFill("solid", colors[vlan])
        cell.font = bold
        cell.alignment = center
 
        # Column headers
        headers = ["S.No", "IP Address", "Allocation", "Device"]
        for i, h in enumerate(headers):
            c = ws.cell(row=row+1, column=col+i, value=h)
            c.fill = PatternFill("solid", "FFFF00")
            c.font = bold
            c.alignment = center
            c.border = border
 
        col += 5  # 4 data columns + 1 empty gap (except after last VLAN)
 
    # Adjust last gap (no empty column after VLAN 13)
    col -= 1
    row += 2
 
    # === Fill 62 IPs per VLAN ===
    all_ips = [list(subnet.hosts()) for subnet in vlan_subnets]
 
    for ip_idx in range(62):
        current_row = row + ip_idx
        col = 1
        for vlan_idx, vlan in enumerate([10, 11, 12, 13]):
            ip = all_ips[vlan_idx][ip_idx]
            pool = vlan_info[vlan]
 
            ws.cell(current_row, col, ip_idx + 1).alignment = center
            ws.cell(current_row, col + 1, str(ip))
 
            if ip_idx == 0:
                alloc = "Gateway IP"        
            elif ip_idx < 10:
                alloc = "Reserved IP"
            elif vlan == 12 and ip_idx == 61:
                alloc = "UPS"
            else:
                alloc = pool
            ws.cell(current_row, col + 2, alloc)
 
            col += 5  # move to next VLAN block (skip empty column)
 
    # === Column widths ===
    widths = [8, 18, 18, 30, 3, 8, 18, 18, 30, 3, 8, 18, 18, 30, 3, 8, 18, 18, 30]  # 19 columns
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w
 
# =============================================
# 4. SAVE
# =============================================
wb.save("Munger.xlsx")
print("SUCCESS! File created: Gaya_75_Locations_HORIZONTAL_WITH_GAP.xlsx")
print("   → One empty column between each VLAN block")
print("   → All 62 usable IPs fully displayed")
