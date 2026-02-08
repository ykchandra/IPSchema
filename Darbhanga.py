# ## Saharsa –
import ipaddress
import re
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter
 
# =============================================
# 1. DATA (locations + allocations)
# =============================================
locations = [
    "Delhi More over bridge both side",
    "Airport gate no. 02",
    "Exit gate of bus stand",
    "Bajar samiti gate",
    "Gehumi over bridge ke niche",
    "Shivdhara chowk",
    "Bagh more",
    "Bhandar chowk",
    "Alinagar mahindra show room",
    "Kadirabad",
    "Hasan Chowk",
    "Shastri Chowk",
    "C.M Arts college more",
    "Saednagar kali mandir Chowk",
    "Bakarganj chowk near Radha krishna mandir",
    "Urdu bajar chowk",
    "Bhigo Chowk Shiv Mandir",
    "Khan Chowk",
    "Near K.S College",
    "Rose Public Bangali Tola More",
    "Balbhadrapur More",
    "New Khajasaray Chowk",
    "Near Railway Gumati Pandasaray Mahavir Mandir",
    "Balbhadrapur Navtoli More 22 no. Railway Gumati jane vali Road",
    "Ganj Chowk",
    "Jail Kona Ahilla Jane vali Road",
    "Ekmi Chowk",
    "Alpatti Chowk",
    "Donar Chowk",
    "Hospital Road",
    "Shastri Chowk",
    "Aazad Chowk",
    "Bhatiyari Sarai",
    "Marwadi College",
    "Mirjapur Chowk More",
    "Khankah Chowk",
    "Darbhanga Tower Chowk",
    "Income tax Chauraha",
    "Lohia Chowk",
    "ICCC"
]

allocation_counts = [
    5, 2, 3, 2, 2, 3, 5, 6, 2, 2, 2, 2, 2, 2, 2, 2, 2, 2, 2, 2, 2, 2, 3, 2, 3, 2, 6, 2, 6, 2, 2, 2, 2, 2, 6, 2, 3, 2, 4, 3
]
vlan_info = {10: "Camera_Pool", 11: "ATCS_ECB_Pool", 12: "UPS_SW_Pool", 13: "VMD_Radar_Pool"}
colors    = {10: "CFE2F3", 11: "D9EAD3", 12: "F4CCCC", 13: "FFE599"}
 
# =============================================
# 2. STYLES
# =============================================
bold = Font(bold=True)
center = Alignment(horizontal="center", vertical="center")
border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
def safe_sheet_name(name):
    name = re.sub(r'[\\/*?:\[\]]', '-', name)
    return name[:31]
# =============================================
# 3. CREATE WORKBOOK
# =============================================
wb = Workbook()
wb.remove(wb.active)
base = ipaddress.IPv4Network("10.10.128.0/26")
subnet_counter = 0
 
for loc_idx, location in enumerate(locations):
   
    ws = wb.create_sheet(title=safe_sheet_name(location))
    switches = allocation_counts[loc_idx]
    row = 1
 
    # === Title ===
    ws.merge_cells('A1:S1')  # now 19 columns
    ws['A1'] = location
    ws['A1'].fill = PatternFill("solid", "1F4E79")
    ws['A1'].font = Font(color="FFFFFF", bold=True, size=14)
    ws['A1'].alignment = center
    row += 2
 
    # === VLAN Headers + Empty Gap ===
    col = 1
    vlan_subnets = []
    for vlan in [10, 11, 12, 13]:
        net_int = int(base.network_address) + subnet_counter * 64
        subnet = ipaddress.IPv4Network((net_int, 26))
        vlan_subnets.append(subnet)
        subnet_counter += 1
 
        # VLAN Title (merged over 4 columns)
        ws.merge_cells(start_row=row, start_column=col, end_row=row, end_column=col+3)
        cell = ws.cell(row=row, column=col, value=f"VLAN {vlan} – {vlan_info[vlan]} – {subnet}")
        cell.fill = PatternFill("solid", colors[vlan])
        cell.font = bold
        cell.alignment = center
 
        # Column headers
        headers = ["S.No", "IP Address", "Allocation", "Device"]
        for i, h in enumerate(headers):
            c = ws.cell(row=row+1, column=col+i, value=h)
            c.fill = PatternFill("solid", "FFFF00")
            c.font = bold
            c.alignment = center
            c.border = border
 
        col += 5  # 4 data columns + 1 empty gap (except after last VLAN)
 
    # Adjust last gap (no empty column after VLAN 13)
    col -= 1
    row += 2
 
    # === Fill 62 IPs per VLAN ===
    all_ips = [list(subnet.hosts()) for subnet in vlan_subnets]
 
    for ip_idx in range(62):
        current_row = row + ip_idx
        col = 1
        for vlan_idx, vlan in enumerate([10, 11, 12, 13]):
            ip = all_ips[vlan_idx][ip_idx]
            pool = vlan_info[vlan]
 
            ws.cell(current_row, col, ip_idx + 1).alignment = center
            ws.cell(current_row, col + 1, str(ip))
 
            if ip_idx == 0:
                alloc = "Gateway IP"        
            elif ip_idx < 10:
                alloc = "Reserved IP"
            elif vlan == 12 and ip_idx == 61:
                alloc = "UPS"
            else:
                alloc = pool
            ws.cell(current_row, col + 2, alloc)
 
            col += 5  # move to next VLAN block (skip empty column)
 
    # === Column widths ===
    widths = [8, 18, 18, 30, 3, 8, 18, 18, 30, 3, 8, 18, 18, 30, 3, 8, 18, 18, 30]  # 19 columns
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w
 
# =============================================
# 4. SAVE
# =============================================
wb.save("Darbhanga.xlsx")
print("SUCCESS! File created: Darbhanga_75_Locations_HORIZONTAL_WITH_GAP.xlsx")
print("   → One empty column between each VLAN block")
print("   → All 62 usable IPs fully displayed")
 
###########################################################################################################
