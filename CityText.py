import ipaddress
import os
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter
 
# =============================================
# 1. DATA (locations + allocations)
# =============================================
locations = [
    "Nagar Thanna Chowk", "Nagar Palika Chowk", "Yogniya Kothi Mod",
    "Sada Over Bridge at west side","Sada Over Bridge ke niche at fish market tri road",
    "Mona Chowk","Mewalal Chowk","Gandhi Chowk","Kathari Bagh tri chowk",
    "Khanua Nalla","Sahebganj Chowk","Nalla No 2- Sonarpatti Road","Mehmood Chowk",
    "Bhikari Chowk","Nehru Chowk","Shishu Park main road",
    "Sadar Hospital Chowk near Mazaar","Daroga Rai Chowk",
    "Bahart Milaap Chowk","Bhagwaan Bazar Station Chowk",
    "Bhagwaan Bazar Thanna Chowk","Near Gudari Bazar Masjid at Chaar Muhani",
    "Barahampur Chowk","Shayamchak","Koniyaa Mayi mandir Chowk",
    "Rajendra Sarovaar Chowk","Mohaya Overbridge both side- Entry",
    "Mohaya Overbridge both side- Exit","Mohaya Overbridge ke Niche (New Bypass )",
    "Novajee Tolla Chowk","Bazar Samiti Chowk","Methvalliya Chowk","Mathiya Chowk",
    "Umdha chowk","Sadda Dalla Overbridge at North side","Chanchora Bazar",
    "Near Magaiedeh Railway Dalla 51","Jhanga Chowk (Near Doriganj Pull)",
    "Kachehri station mod","Kashi Bazar Chowk","Prabhu Nath Nagar PCC mod",
    "Bishnpura Mohiya mode","Mukrera mod","Bada Telpa","Police Line Gate",
    "Lower road nayi Bazaar mod (teen mohani)","Lal Bazaar",
    "Garkha dhala rathore tola","Gheghtha Bazar","Meera Musheri Mod",
    "Dwarka Dheesh Mod","Turkaulia","Near Enayi Hanumaan Mandir",
    "State Bank of India Near Pankaj Cinema","Bazar Samiti Main Gate",
    "Chota Telpa Chowk"
]
 
allocation_counts = [
    2,5,2,2,2,2,2,2,2,2,
    2,2,2,2,2,3,2,4,2,4,
    2,2,4,2,3,4,2,2,2,4,
    2,2,2,2,2,2,2,2,3,2,
    2,2,2,2,2,2,3,2,2,2,
    2,2,2,2,3,2
]
 
vlan_info = {10: "Camera_Pool", 11: "ATCS_ECB_Pool", 12: "UPS_SW_Pool", 13: "VMD_Radar_Pool"}
colors    = {10: "CFE2F3", 11: "D9EAD3", 12: "F4CCCC", 13: "FFE599"}
 
# =============================================
# 2. STYLES
# =============================================
bold = Font(bold=True)
center = Alignment(horizontal="center", vertical="center")
border = Border(left=Side(style='thin'), right=Side(style='thin'),
                top=Side(style='thin'), bottom=Side(style='thin'))
 
# =============================================
# 3. CREATE WORKBOOK
# =============================================
wb = Workbook()
wb.remove(wb.active)
 
# Base network for all VLANs
base = ipaddress.IPv4Network("10.10.128.0/26")
subnet_counter = 0
 
# Create main folder for Switch TXT files
main_folder = "Switch_Configs_Chapra"
os.makedirs(main_folder, exist_ok=True)
 
# =============================================
# PROCESS EACH LOCATION
# =============================================
for loc_idx, location in enumerate(locations):
 
    # -------------------------------
    # CREATE EXCEL SHEET
    # -------------------------------
    ws = wb.create_sheet(title=location[:31])
    switches = allocation_counts[loc_idx]
    row = 1
 
    ws.merge_cells('A1:S1')
    ws['A1'] = location
    ws['A1'].fill = PatternFill("solid", "1F4E79")
    ws['A1'].font = Font(color="FFFFFF", bold=True, size=14)
    ws['A1'].alignment = center
    row += 2
 
    col = 1
    vlan_subnets = []
 
    # -------------------------------
    # GENERATE SUBNETS FOR 4 VLANs
    # -------------------------------
    for vlan in [10,11,12,13]:
        net_int = int(base.network_address) + subnet_counter * 64
        subnet = ipaddress.IPv4Network((net_int, 26))
        vlan_subnets.append(subnet)
        subnet_counter += 1
 
        ws.merge_cells(start_row=row, start_column=col, end_row=row, end_column=col+3)
        hdr = ws.cell(row=row, column=col,
                      value=f"VLAN {vlan} – {vlan_info[vlan]} – {subnet}")
        hdr.fill = PatternFill("solid", colors[vlan])
        hdr.font = bold
        hdr.alignment = center
 
        headers = ["S.No", "IP Address", "Allocation", "Device"]
        for i, h in enumerate(headers):
            c = ws.cell(row=row+1, column=col+i, value=h)
            c.fill = PatternFill("solid", "FFFF00")
            c.font = bold
            c.alignment = center
            c.border = border
 
        col += 5
 
    col -= 1
    row += 2
 
    # -------------------------------
    # FILL 62 IPs IN EXCEL
    # -------------------------------
    all_ips = [list(subnet.hosts()) for subnet in vlan_subnets]
 
    for ip_idx in range(62):
        current_row = row + ip_idx
        col = 1
 
        for vlan_idx, vlan in enumerate([10,11,12,13]):
            ip = all_ips[vlan_idx][ip_idx]
            pool = vlan_info[vlan]
 
            ws.cell(current_row, col, ip_idx+1).alignment = center
            ws.cell(current_row, col+1, str(ip))
 
            if ip_idx == 0:
                alloc = "Gateway IP"
            elif vlan == 12 and ip_idx <= allocation_counts[loc_idx]:
                alloc = location + " Switch IP"
            elif ip_idx < 10:
                alloc = "Reserved IP"
            elif ip_idx == 61:
                alloc = "UPS"
            else:
                alloc = pool
 
            ws.cell(current_row, col+2, alloc)
 
            col += 5
 
    widths = [8,18,18,30,3,8,18,18,30,3,8,18,40,30,3,8,18,18,30]
    for i,w in enumerate(widths,1):
        ws.column_dimensions[get_column_letter(i)].width = w
 
    # -------------------------------
    # CREATE SWITCH CONFIG FILES
    # -------------------------------
    loc_folder = os.path.join(main_folder, location.replace(" ", "_"))
    os.makedirs(loc_folder, exist_ok=True)
 
    vlan12_subnet = vlan_subnets[2]     # 3rd subnet = VLAN 12
    hosts = list(vlan12_subnet.hosts())
    gateway_ip = hosts[0]
    switch_ips = hosts[1:]
 
    for sw in range(switches):
 
        sw_hostname = f"{location.replace(' ', '_')}_SW{sw+1}"
        sw_ip = switch_ips[sw]
 
        config = f"""
conf t
hostname {sw_hostname}
enable secret S@fEC!tY#2025
username network password S@fEC!tY#2025
!
interface vlan1
no ip address
exit
 
vlan 10
vlan 11
vlan 12
vlan 13
 
interface vlan 12
ip address {sw_ip} 255.255.255.192
no shutdown
!
ip default-gateway {gateway_ip}
 
interface range gi 1/1-2
switchport mode trunk
no shutdown
 
interface range gi 1/3-12
switchport mode access
switchport access vlan XXXX
exit
 
snmp-server host {gateway_ip} #0Tl@d1e5
snmp-server community #0Tl@d1e5 RW
snmp-server enable traps
 
line vty 0 4
password S@fEC!tY#2025
login
!
end
write
"""
 
        file_path = os.path.join(loc_folder, f"{sw_hostname}.txt")
        with open(file_path, "w") as f:
            f.write(config.strip())
 
# =============================================
# SAVE EXCEL
# =============================================
wb.save("Chapratest.xlsx")
print("SUCCESS: Excel + Switch configs generated!")